import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


DARK_PURPLE = "2D1B69"
ORANGE = "E8824A"
LIGHT_PURPLE = "EDE8F5"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"

COLUMNS = ["Date", "Vendor", "Category", "Amount", "Payment Method", "Notes", "Source File"]
KEYS = ["date", "vendor", "category", "amount", "payment_method", "notes", "source_file"]


def to_excel(expenses: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.merge_cells("A1:G1")
    ws["A1"] = "Future-Forward Concepts LLC — Expense Tracker"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws["A1"].fill = PatternFill("solid", start_color=DARK_PURPLE, fgColor=DARK_PURPLE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    for col, header in enumerate(COLUMNS, 1):
        c = ws.cell(row=2, column=col, value=header)
        c.font = Font(name="Arial", bold=True, color=WHITE)
        c.fill = PatternFill("solid", start_color=ORANGE, fgColor=ORANGE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[2].height = 18

    # Data rows
    for i, exp in enumerate(expenses):
        r = i + 3
        fill_color = LIGHT_PURPLE if i % 2 == 0 else LIGHT_GRAY
        for col, key in enumerate(KEYS, 1):
            val = exp.get(key, "")
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = PatternFill("solid", start_color=fill_color, fgColor=fill_color)
            c.border = border
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(col == 6))
            if col == 4 and val:
                c.number_format = '$#,##0.00'
                c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 30

    # Total row
    total_row = len(expenses) + 3
    ws.merge_cells(f"A{total_row}:C{total_row}")
    lbl = ws.cell(row=total_row, column=1, value="TOTAL")
    lbl.font = Font(name="Arial", bold=True, color=WHITE)
    lbl.fill = PatternFill("solid", start_color=DARK_PURPLE, fgColor=DARK_PURPLE)
    lbl.alignment = Alignment(horizontal="right", vertical="center")
    lbl.border = border

    tot = ws.cell(row=total_row, column=4, value=f"=SUM(D3:D{total_row-1})")
    tot.font = Font(name="Arial", bold=True, color=WHITE)
    tot.fill = PatternFill("solid", start_color=DARK_PURPLE, fgColor=DARK_PURPLE)
    tot.number_format = '$#,##0.00'
    tot.alignment = Alignment(horizontal="right", vertical="center")
    tot.border = border

    for col in range(5, 8):
        c = ws.cell(row=total_row, column=col, value="")
        c.fill = PatternFill("solid", start_color=DARK_PURPLE, fgColor=DARK_PURPLE)
        c.border = border

    ws.row_dimensions[total_row].height = 18

    # Column widths
    widths = [12, 30, 16, 12, 18, 40, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_csv(expenses: list[dict]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=KEYS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(expenses)
    return buf.getvalue().encode("utf-8")
